# -*- coding: utf-8 -*-
import io
import json
import re
import tempfile
from pathlib import Path

import openpyxl
from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel

from config import MODEL_CHAT
from llm_client import get_llm_client

router = APIRouter(prefix="/api/audit")


# ── 分类体系常量 ───────────────────────────────────────────────────

CATEGORY_TAXONOMY: dict[str, list[str]] = {
    "公司治理": ["三重一大决策", "董事会管理", "股东会管理", "会议管理", "其他"],
    "合同及法律合规管理": ["合同审核及签署", "合同执行", "诉讼管理", "授权管理", "知识产权管理", "其他"],
    "采购管理": ["采购方式", "采购评审", "供应商管理", "采购文档管理", "其他"],
    "营销管理": ["代理人管理", "销售管理", "大客户管理", "团队管理", "常旅客管理", "知音商城管理", "品牌管理", "其他"],
    "人力资源管理": ["人员招聘", "绩效考核", "薪酬福利", "考勤管理", "培训管理", "岗位与人员配置", "离职管理", "领导人员履职待遇", "其他"],
    "财务管理": ["预算管理", "银行账户管理", "成本费用管理", "资金管理", "往来账款管理", "会计核算管理", "保险及索赔管理", "担保管理", "税务管理", "优惠政策使用管理", "其他"],
    "资产管理": ["固定资产管理", "存货管理", "低值易耗品管理", "无形资产管理", "资产权证管理", "其他"],
    "信息系统管理": ["系统功能开发管理", "系统账号管理", "系统安全管理", "系统应用管理", "其他"],
    "工程项目管理": ["工程项目工期管理", "工程项目招标管理", "工程项目洽商变更", "施工过程管理", "工程项目验收", "工程项目竣工结决算", "其他"],
    "安全管理": ["安全事件", "安全与质量考核", "空防、消防、地面安全管理", "应急管理", "其他"],
    "内部控制管理": ["评价管理", "内部控制手册建设", "风险识别与管理", "规章制度审核", "其他"],
    "行政管理（其他）": ["中央八项规定精神", "档案管理", "证照管理", "礼品管理", "印章管理", "免折票管理", "审计整改", "对外捐赠", "企业文化"],
    "其他": ["其他"],
}


# ── 数据模型 ──────────────────────────────────────────────────────


class AuditRow(BaseModel):
    seq: int
    issue: str
    description: str
    category_l1: str = ""
    category_l2: str = ""
    domain: str = ""


class DownloadRequest(BaseModel):
    rows: list[AuditRow]
    original_filename: str = "审计问题分析结果"


# ── 工具函数 ──────────────────────────────────────────────────────


def _find_header_row(ws) -> int | None:
    """找到表头行：某单元格值精确等于「发现问题」的行（1-based）。
    使用精确匹配避免把含「发现问题」子串的标题行（如「审计发现问题汇总表」）误判为表头。
    """
    for row_idx in range(1, min(ws.max_row + 1, 10)):
        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(row_idx, col_idx).value
            if cell_val and str(cell_val).strip() == "发现问题":
                return row_idx
    return None


def _col_index(ws, header_row: int, keyword: str) -> int | None:
    """在标题行中找含 keyword 的列号（1-based）。"""
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(header_row, col_idx).value
        if val and keyword in str(val):
            return col_idx
    return None


def _extract_rows(wb: openpyxl.Workbook) -> list[dict]:
    """提取第一个 sheet 中的审计发现行数据。"""
    ws = wb.active
    header_row = _find_header_row(ws)
    if header_row is None:
        raise ValueError("未找到包含「发现问题」的标题行，请检查 Excel 格式。")

    seq_col = _col_index(ws, header_row, "序号") or 1
    issue_col = _col_index(ws, header_row, "发现问题")
    desc_col = _col_index(ws, header_row, "问题描述")

    if issue_col is None:
        raise ValueError("未找到「发现问题」列。")

    rows = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        seq_val = ws.cell(row_idx, seq_col).value
        issue_val = ws.cell(row_idx, issue_col).value
        desc_val = ws.cell(row_idx, desc_col).value if desc_col else ""

        if not issue_val:
            continue

        try:
            seq = int(seq_val) if seq_val is not None else row_idx - header_row
        except (ValueError, TypeError):
            seq = row_idx - header_row

        rows.append({
            "seq": seq,
            "issue": str(issue_val).strip(),
            "description": str(desc_val).strip() if desc_val else "",
        })

    return rows


def _build_prompt(rows: list[dict], domains: list[str]) -> str:
    rows_json = json.dumps(
        [{"序号": r["seq"], "发现问题": r["issue"], "问题描述": r["description"][:200]}
         for r in rows],
        ensure_ascii=False,
        indent=2,
    )

    # 构建一级→二级对应关系说明
    taxonomy_lines = "\n".join(
        f"- {l1}：{'/ '.join(l2_list)}"
        for l1, l2_list in CATEGORY_TAXONOMY.items()
    )

    domain_lines = "\n".join(f"- {d}" for d in domains)

    return f"""你是企业内部审计专家。请对以下审计发现问题进行分类，共两个独立维度，分别判断。

【维度一：问题类别（两级）】
请先选择最匹配的一级类别，再在该一级类别下选择最匹配的二级子类。

一级类别及其对应二级子类：
{taxonomy_lines}

【维度二：业务领域】
描述问题所属的业务板块，从以下选项中选一个最匹配的：
{domain_lines}

【发现问题列表（JSON）】：
{rows_json}

请严格按如下格式返回，只输出JSON数组，不含任何其他文字：
[{{"序号": 1, "问题类别一级": "...", "问题类别二级": "...", "业务领域": "..."}}, ...]"""


def _parse_llm_output(text: str, rows: list[dict]) -> list[dict]:
    """从 LLM 输出中提取 JSON 数组，补全缺失行。"""
    match = re.search(r'\[.*\]', text, re.DOTALL)
    if not match:
        raise ValueError(f"LLM 返回格式异常，无法解析 JSON：{text[:200]}")

    parsed = json.loads(match.group())
    result_map = {item["序号"]: item for item in parsed}

    result = []
    for r in rows:
        classified = result_map.get(r["seq"], {})
        result.append({
            **r,
            "category_l1": classified.get("问题类别一级", ""),
            "category_l2": classified.get("问题类别二级", ""),
            "domain": classified.get("业务领域", ""),
        })
    return result


# ── 路由 ──────────────────────────────────────────────────────────


@router.post("/analyze")
async def analyze_audit(
    file: UploadFile = File(...),
    domains: str = Form('["物业租赁","酒店公寓","工程领域","资产处置","历史遗留问题"]'),
):
    try:
        doms = json.loads(domains)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"业务领域参数格式错误：{e}")

    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        rows = _extract_rows(wb)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel 解析失败：{e}")

    if not rows:
        raise HTTPException(status_code=400, detail="Excel 中未找到有效数据行。")

    prompt = _build_prompt(rows, doms)
    try:
        client = get_llm_client()
        resp = client.chat.completions.create(
            model=MODEL_CHAT,
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
        )
        llm_text = resp.choices[0].message.content or ""
        classified_rows = _parse_llm_output(llm_text, rows)
    except ValueError as e:
        raise HTTPException(status_code=502, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"LLM 调用失败：{e}")

    return {"rows": classified_rows, "total": len(classified_rows)}


@router.post("/download")
async def download_audit_excel(req: DownloadRequest):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "审计问题分析"

    # 样式定义
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_align = Alignment(wrap_text=True, vertical="top")

    headers = ["序号", "发现问题", "问题描述", "问题类别一级", "问题类别二级", "业务领域"]
    col_widths = [8, 30, 50, 18, 20, 14]

    # 写标题行
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 20

    # 分类列填充色
    cat_l1_fill = PatternFill("solid", fgColor="EBF5FB")
    cat_l2_fill = PatternFill("solid", fgColor="D6EAF8")
    dom_fill = PatternFill("solid", fgColor="E9F7EF")

    for row_idx, row in enumerate(req.rows, start=2):
        ws.cell(row_idx, 1, row.seq).alignment = center_align
        ws.cell(row_idx, 2, row.issue).alignment = wrap_align
        ws.cell(row_idx, 3, row.description).alignment = wrap_align
        l1_cell = ws.cell(row_idx, 4, row.category_l1)
        l1_cell.alignment = center_align
        l1_cell.fill = cat_l1_fill
        l2_cell = ws.cell(row_idx, 5, row.category_l2)
        l2_cell.alignment = center_align
        l2_cell.fill = cat_l2_fill
        dom_cell = ws.cell(row_idx, 6, row.domain)
        dom_cell.alignment = center_align
        dom_cell.fill = dom_fill

    # 冻结标题行
    ws.freeze_panes = "A2"

    # 写入临时文件并返回
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    wb.save(tmp_path)

    safe_name = req.original_filename.replace("/", "_").replace("\\", "_")
    return FileResponse(
        path=tmp_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{safe_name}_分类结果.xlsx",
    )
