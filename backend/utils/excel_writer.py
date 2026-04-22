import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from config import DATA_ROOT


EXCEL_PATH = os.path.join(DATA_ROOT, "培训统计表.xlsx")

HEADERS = ["序号", "培训日期", "培训主题", "培训地点", "主办部门", "参与人数", "培训类别", "归档路径", "录入时间"]


def init_excel():
    """初始化 Excel 文件（如果不存在则创建）"""
    if os.path.exists(EXCEL_PATH):
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "培训记录"

    # 写表头
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 设置列宽
    col_widths = [6, 14, 30, 20, 16, 10, 14, 50, 20]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 22
    wb.save(EXCEL_PATH)


def append_record(
    date: str,
    topic: str,
    location: str,
    department: str,
    count: int,
    category: str,
    archive_path: str,
):
    """在 Excel 中追加一条培训记录"""
    init_excel()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    next_row = ws.max_row + 1
    seq = next_row - 1  # 序号（去掉表头行）

    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    values = [seq, date, topic, location, department, count, category, archive_path, now]
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=next_row, column=col, value=value)
        cell.alignment = Alignment(vertical="center")
        # 奇偶行背景色
        if next_row % 2 == 0:
            cell.fill = PatternFill(fill_type="solid", fgColor="DCE6F1")

    wb.save(EXCEL_PATH)
    return EXCEL_PATH
