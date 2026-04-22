"""
Excel 写入模块
按照"诉讼案件台账AI"样式生成台账，支持多审级合并单元格
"""

from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Side, Font, PatternFill
)
from openpyxl.utils import get_column_letter


# ── 样式常量 ──────────────────────────────────────────────────

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FILL  = PatternFill("solid", fgColor="BDD7EE")   # 蓝色表头
TITLE_FONT   = Font(name="微软雅黑", bold=True, size=12)
HEADER_FONT  = Font(name="微软雅黑", bold=True, size=9)
DATA_FONT    = Font(name="微软雅黑", size=9)

WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")

# 17 列宽度（A-Q）
COL_WIDTHS = [6, 12, 28, 16, 30, 12, 12, 45, 14, 45, 8, 14, 30, 12, 12, 12, 25]


def _style_cell(cell, font=None, fill=None, alignment=None, border=True):
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border    = BORDER


def _write_header(ws):
    """写标题行、来源行、字段名行（共 3 行）"""
    # Row 1: 标题
    ws.merge_cells("A1:Q1")
    c = ws["A1"]
    c.value     = "（中国航空集团建设开发有限公司）法律纠纷案件明细表"
    c.font      = TITLE_FONT
    c.alignment = CENTER_ALIGN
    c.fill      = PatternFill("solid", fgColor="DEEAF1")
    ws.row_dimensions[1].height = 22

    # Row 2: 来源说明
    labels = {
        "A2": "来源：",
        "B2": "起诉状",
        "I2": "判决书",
        "L2": "强制执行申请书",
        "M2": "执行裁定(如有)",
    }
    merges2 = ["B2:H2", "I2:K2", "M2:N2"]
    for rng in merges2:
        ws.merge_cells(rng)
    for addr, val in labels.items():
        c = ws[addr]
        c.value     = val
        c.font      = HEADER_FONT
        c.fill      = HEADER_FILL
        c.alignment = CENTER_ALIGN
        c.border    = BORDER
    # 给 Row2 其余单元格加边框和填充
    for col in range(1, 18):
        c = ws.cell(row=2, column=col)
        if not c.value:
            c.fill   = HEADER_FILL
            c.border = BORDER
    ws.row_dimensions[2].height = 16

    # Row 3: 字段名
    headers = [
        "序号", "案件发生时间", "案件名称", "案由", "诉讼主体",
        "主诉/被诉", "标的金额（万元）",
        "基本情况（来源于业务单位情况说明及起诉状诉讼请求）",
        "生效判决/裁定日期", "处理结果", "审级",
        "强制执行时间", "执行（判决履行）情况（人工填写项）",
        "执行回款金额\n（万元，人工填写项）",
        "挽回损失（人工填写项）", "支出费用（人工填写项）",
        "服务律所（人工填写项）",
    ]
    ws.merge_cells("J3:K3")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col)
        if col != 11:  # K3 被 J3 合并，跳过
            c.value = h
        _style_cell(c, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER_ALIGN)
    ws.row_dimensions[3].height = 36


def _write_case(ws, row_start: int, seq: int, case: dict):
    """
    写一个案件的所有行（按审级数量决定行数）。
    返回写完后下一个案件的起始行号。
    """
    stages = case.get("stages") or []
    n_rows = max(len(stages), 1)  # 至少占 1 行
    row_end = row_start + n_rows - 1

    # ── 需要合并的列（A-I, L-Q）──────────────────────────────
    merge_cols = list(range(1, 10)) + list(range(12, 18))  # A-I, L-Q

    # 填写案件级别字段
    case_values = [
        seq,
        case.get("案件发生时间"),
        case.get("案件名称", ""),
        case.get("案由", ""),
        case.get("诉讼主体", ""),
        case.get("主诉被诉", ""),
        case.get("标的金额"),
        case.get("基本情况", ""),
        case.get("生效判决日期"),
        None,  # J: 处理结果（per-stage）
        None,  # K: 审级（per-stage）
        case.get("强制执行时间"),
        "",    # M: 执行情况（人工）
        None,  # N: 执行回款（人工）
        None,  # O: 挽回损失（人工）
        None,  # P: 支出费用（人工）
        case.get("服务律所", ""),
    ]

    # 写第一行（合并列的值写在这里）
    for col, val in enumerate(case_values, 1):
        c = ws.cell(row=row_start, column=col)
        c.value = val
        _style_cell(c, font=DATA_FONT, alignment=WRAP_ALIGN)
        # 序号居中
        if col == 1:
            c.alignment = CENTER_ALIGN

    # 合并 & 补边框
    if n_rows > 1:
        for col in merge_cols:
            ws.merge_cells(
                start_row=row_start, start_column=col,
                end_row=row_end,     end_column=col
            )
        # 给合并区域内每个子单元格加边框（openpyxl 合并后边框只在首格）
        for r in range(row_start + 1, row_end + 1):
            for col in merge_cols:
                ws.cell(row=r, column=col).border = BORDER

    # ── 写各审级处理结果 ──────────────────────────────────────
    for i, stage in enumerate(stages):
        r = row_start + i
        # J: 处理结果
        cj = ws.cell(row=r, column=10)
        cj.value     = stage.get("处理结果", "")
        cj.font      = DATA_FONT
        cj.alignment = WRAP_ALIGN
        cj.border    = BORDER
        # K: 审级
        ck = ws.cell(row=r, column=11)
        ck.value     = stage.get("审级", "")
        ck.font      = DATA_FONT
        ck.alignment = CENTER_ALIGN
        ck.border    = BORDER

    # 如果没有审级数据，J/K 也要有边框
    if not stages:
        for col in [10, 11]:
            c = ws.cell(row=row_start, column=col)
            c.border = BORDER

    # 行高：按内容估算（每 50 字符约 1 行，每行 15pt）
    for r in range(row_start, row_end + 1):
        ws.row_dimensions[r].height = 60

    return row_end + 1


def write_ledger(cases: list[dict], output_path: str):
    """
    将案件数据列表写入 Excel 台账。

    参数:
        cases: build_case_data() 返回的案件数据列表
        output_path: 输出文件路径
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "诉讼案件台账"

    # 设置列宽
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 冻结窗格（从第4行开始滚动）
    ws.freeze_panes = "A4"

    # 写表头
    _write_header(ws)

    # 写数据
    current_row = 4
    for seq, case in enumerate(cases, 1):
        current_row = _write_case(ws, current_row, seq, case)

    # 保存
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"台账已保存: {output_path}")


# ── 命令行测试 ────────────────────────────────────────────────

if __name__ == "__main__":
    # 用一条假数据测试格式
    test_cases = [
        {
            "案件名称": "四川中航物业与成都中大中医房屋租赁纠纷案",
            "案件发生时间": "2022-08-05",
            "案由": "房屋租赁合同纠纷",
            "诉讼主体": "原告：四川中航建开物业管理有限责任公司\n被告一：成都中大中医研究院\n被告二：丁正",
            "主诉被诉": "诉讼-主诉",
            "标的金额": 809.72,
            "基本情况": "一审诉讼请求：判决被告立即腾退房屋，并支付逾期占用费807300元及资金占用费5852.93元。",
            "生效判决日期": "2024-09-18",
            "强制执行时间": "2024-08-23",
            "服务律所": "北京德恒（成都）律师事务所郑显芳律师",
            "stages": [
                {"审级": "一审", "处理结果": "武侯区法院2023年11月11日判决：驳回原告四川中航建开物业管理有限公司的诉讼请求。"},
                {"审级": "二审", "处理结果": "成都中院2024年8月6日终审判决：撤销一审判决，判决被告支付房屋占用费并计利息。"},
                {"审级": "再审", "处理结果": "四川高院2024年12月5日裁定：驳回再审申请。"},
            ],
        }
    ]

    out = r"C:\Users\shinh\Documents\GitHub\claude-code\legal_ledger\output\测试台账.xlsx"
    write_ledger(test_cases, out)
