# -*- coding: utf-8 -*-
import io
from typing import Optional
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


def normalize_contract_no(s: str) -> str:
    """标准化合同编号用于模糊匹配（大小写、全角括号、末尾"号"）"""
    s = str(s).strip()
    s = s.lower()
    s = s.replace('（', '(').replace('）', ')')
    if s.endswith('号'):
        s = s[:-1]
    return s.strip()


def clean_header(h) -> str:
    """清洗表头值：去除星号和首尾空格"""
    if h is None:
        return ''
    return str(h).strip().lstrip('*').strip()


def find_header_row_idx(ws) -> int:
    """扫描前6行，返回非空单元格最多的那行索引（1-based）"""
    best_row = 1
    best_count = 0
    for row_idx in range(1, min(7, ws.max_row + 1)):
        count = sum(
            1 for cell in ws[row_idx]
            if cell.value is not None and str(cell.value).strip()
        )
        if count > best_count:
            best_count = count
            best_row = row_idx
    return best_row


def read_file_as_records(file_bytes: bytes, label: str) -> tuple[list[dict], str]:
    """
    读取Excel文件，返回 (records, contract_col_name)。
    records 中每条记录的 key 格式为 "[label] 列名"。
    contract_col_name 是合同编号列在 records 中的 key。
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # 多Sheet时优先选含"汇总"的Sheet，否则取第一个
    sheet_name = wb.sheetnames[0]
    for name in wb.sheetnames:
        if '汇总' in name:
            sheet_name = name
            break
    ws = wb[sheet_name]

    header_row_idx = find_header_row_idx(ws)

    # 提取并清洗表头
    raw_headers = [clean_header(cell.value) for cell in ws[header_row_idx]]

    # 找合同编号列（index）
    contract_col_idx = None
    for i, h in enumerate(raw_headers):
        if '合同编号' in h:
            contract_col_idx = i
            break
    if contract_col_idx is None:
        raise ValueError(f'在"{label}台账"中未找到包含"合同编号"的列，请检查文件格式。')

    # 合计行关键字（任一单元格值完全匹配则跳过整行）
    SUMMARY_KEYWORDS = {'合计行', '合计', '小计', '总计', '合  计', '小  计'}

    # 读数据行，跳过全空行和合计行
    records = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if all(v is None or str(v).strip() == '' for v in row):
            continue
        if any(str(v).strip() in SUMMARY_KEYWORDS for v in row if v is not None):
            continue
        row_dict = {}
        for i, h in enumerate(raw_headers):
            if not h:
                continue
            val = row[i] if i < len(row) else None
            row_dict[f'[{label}] {h}'] = val
        records.append(row_dict)

    contract_col_name = f'[{label}] {raw_headers[contract_col_idx]}'
    return records, contract_col_name


def build_index(records: list[dict], contract_col: str) -> dict[str, list[dict]]:
    """构建 normalized_no -> [rows] 索引"""
    index: dict[str, list[dict]] = {}
    for row in records:
        raw = row.get(contract_col)
        if raw is None or str(raw).strip() == '':
            continue
        key = normalize_contract_no(str(raw))
        index.setdefault(key, []).append(row)
    return index


def determine_status(has_purchase: bool, matched_purchase: bool,
                     has_finance: bool, matched_finance: bool) -> tuple[str, int]:
    """
    返回 (状态文本, 排序优先级)
    优先级: 1=全部匹配, 2=部分匹配, 3=全不匹配
    """
    missing = []
    if has_purchase and not matched_purchase:
        missing.append('采购表')
    if has_finance and not matched_finance:
        missing.append('财务表')

    if not missing:
        return '全部匹配', 1

    total_uploaded = int(has_purchase) + int(has_finance)
    if len(missing) == total_uploaded:
        # 上传的所有辅助表都没匹配上
        return '缺少' + '和'.join(missing) + '内容', 3
    else:
        return '缺少' + '和'.join(missing) + '内容', 2


def merge_ledgers(
    contract_bytes: bytes,
    purchase_bytes: Optional[bytes],
    finance_bytes: Optional[bytes],
) -> tuple[bytes, dict]:
    """
    主合并函数，返回 (Excel bytes, 统计信息字典)
    """
    contract_records, contract_col = read_file_as_records(contract_bytes, '合同')
    purchase_records, purchase_col = (
        read_file_as_records(purchase_bytes, '采购') if purchase_bytes else ([], '')
    )
    finance_records, finance_col = (
        read_file_as_records(finance_bytes, '财务') if finance_bytes else ([], '')
    )

    purchase_index = build_index(purchase_records, purchase_col) if purchase_bytes else {}
    finance_index = build_index(finance_records, finance_col) if finance_bytes else {}

    # 收集各表的所有列名（保持顺序，去重）
    def unique_cols(records: list[dict]) -> list[str]:
        seen: set[str] = set()
        cols: list[str] = []
        for row in records:
            for k in row:
                if k not in seen:
                    seen.add(k)
                    cols.append(k)
        return cols

    contract_cols = unique_cols(contract_records)
    purchase_cols = unique_cols(purchase_records)
    finance_cols = unique_cols(finance_records)

    all_output_cols = ['匹配状态', '合同编号'] + contract_cols + purchase_cols + finance_cols

    # 统计计数
    fully_matched = 0
    partial_matched = 0
    unmatched = 0
    matched_purchase_count = 0
    matched_finance_count = 0

    output_rows: list[dict] = []
    for c_row in contract_records:
        raw_no = c_row.get(contract_col, '')
        norm_no = normalize_contract_no(str(raw_no)) if raw_no else ''

        p_matches = purchase_index.get(norm_no) if purchase_bytes else None
        f_matches = finance_index.get(norm_no) if finance_bytes else None

        matched_purchase = bool(p_matches)
        matched_finance = bool(f_matches)

        status, priority = determine_status(
            has_purchase=bool(purchase_bytes),
            matched_purchase=matched_purchase,
            has_finance=bool(finance_bytes),
            matched_finance=matched_finance,
        )

        if priority == 1:
            fully_matched += 1
        elif priority == 2:
            partial_matched += 1
        else:
            unmatched += 1

        if matched_purchase:
            matched_purchase_count += 1
        if matched_finance:
            matched_finance_count += 1

        p_list = p_matches if matched_purchase else [None]
        f_list = f_matches if matched_finance else [None]

        for p_row in p_list:
            for f_row in f_list:
                out: dict = {
                    '匹配状态': status,
                    '_priority': priority,
                    '合同编号': raw_no,
                }
                out.update(c_row)
                if p_row:
                    out.update(p_row)
                else:
                    for col in purchase_cols:
                        out[col] = None
                if f_row:
                    out.update(f_row)
                else:
                    for col in finance_cols:
                        out[col] = None
                output_rows.append(out)

    # 去重：按输出列内容做完全一致性比对，保留首次出现的行
    seen: set[tuple] = set()
    deduped: list[dict] = []
    for row in output_rows:
        key = tuple(str(row.get(col)) for col in all_output_cols)
        if key not in seen:
            seen.add(key)
            deduped.append(row)
    output_rows = deduped

    # 按优先级稳定排序
    output_rows.sort(key=lambda r: r.get('_priority', 99))

    excel_bytes = _write_excel(output_rows, all_output_cols)

    stats = {
        'total_contract': len(contract_records),
        'matched_purchase': matched_purchase_count,
        'matched_finance': matched_finance_count,
        'fully_matched': fully_matched,
        'partial_matched': partial_matched,
        'unmatched': unmatched,
    }
    return excel_bytes, stats


def _write_excel(rows: list[dict], columns: list[str]) -> bytes:
    """生成带格式的Excel，返回bytes"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '合并台账'

    fill_key = PatternFill('solid', fgColor='FFEB9C')       # 浅黄（状态/编号列表头）
    fill_contract = PatternFill('solid', fgColor='BDD7EE')  # 浅蓝（合同系统）
    fill_purchase = PatternFill('solid', fgColor='C6EFCE')  # 浅绿（采购系统）
    fill_finance = PatternFill('solid', fgColor='FCE4D6')   # 浅橙（财务系统）
    fill_ok = PatternFill('solid', fgColor='C6EFCE')        # 绿（全匹配）
    fill_partial = PatternFill('solid', fgColor='FFEB9C')   # 黄（部分匹配）
    fill_none = PatternFill('solid', fgColor='FFC7CE')      # 红（全不匹配）
    bold = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 写表头
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = bold
        cell.alignment = center
        if col_name in ('匹配状态', '合同编号'):
            cell.fill = fill_key
        elif col_name.startswith('[合同]'):
            cell.fill = fill_contract
        elif col_name.startswith('[采购]'):
            cell.fill = fill_purchase
        elif col_name.startswith('[财务]'):
            cell.fill = fill_finance
        else:
            cell.fill = fill_key

    # 写数据行
    for row_idx, row in enumerate(rows, 2):
        priority = row.get('_priority', 2)
        for col_idx, col_name in enumerate(columns, 1):
            val = row.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_name == '匹配状态':
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if priority == 1:
                    cell.fill = fill_ok
                elif priority == 2:
                    cell.fill = fill_partial
                else:
                    cell.fill = fill_none

    # 动态调整列宽（上限40）
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(col_name)
        for row in rows:
            val = row.get(col_name)
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
